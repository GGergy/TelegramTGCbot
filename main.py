import datetime as dtm
import telebot
import json
from telebot import types
from openpyxl import load_workbook
from config import TOKEN


VERSION = "0.2.0 stable by Gergy"
users = {}
bot = telebot.TeleBot(TOKEN)
admins_id = [1694307474, 5468165968]
languages = ['Русский', 'English']
languages_emojies = {"Русский": "🇷🇺", "English": "🇬🇧", 'Poland': "🇵🇱"}
wb = load_workbook('data.xlsx')

try:
    with open("passwords.json", "r") as file:
        passwords = json.load(file)
except:
    with open('passwords.json', 'w') as f:
        f.write('{}')


class User:
    def __init__(self, id) -> None:
        self.id = str(id)
        self.verify = False
        self.is_online = False
        self.language = None
        self.roots = None

    def upload_info(self):
        wb.create_sheet(str(self.id))
        sheet = wb[self.id]
        sheet['A1'].value = self.name
        sheet['A2'].value = self.nname
        sheet['A3'].value = self.gender
        sheet['A4'].value = self.region
        sheet['A5'].value = self.tg_name
        sheet['A6'].value = self.date
        wb.save('data.xlsx')
        print(1)
        self.verify = True

    def load_info(self):
        sheet = wb[self.id]
        self.name = sheet['A1'].value
        self.gender = sheet['A3'].value
        self.nname = sheet['A2'].value
        self.region = sheet['A4'].value
        self.tg_name = sheet['A5'].value
        self.date = sheet['A6'].value
        self.roots = [el for el in [elem.value for elem in sheet['B']]]
        self.is_online = True

    def __str__(self):
        return f'{self.name} {self.nname} {self.gender} {self.region} {self.roots}'


@bot.message_handler(commands=['start'])
def start(message):
    bot.send_message(message.chat.id, f'Current version: {VERSION}. For questions - @TGurchinC or @AsusROG_fan')
    bot.send_message(message.chat.id, f'Bot developming in progress, if bot dont answer please wait')
    if message.chat.id not in users or type(users[message.chat.id]) != 'User':
        users[message.chat.id] = User(message.chat.id)
    markup = types.InlineKeyboardMarkup()
    for lg in languages:
        b1 = types.InlineKeyboardButton(languages_emojies[lg] + lg + languages_emojies[lg], callback_data=lg)
        markup.row(b1)
    bot.send_message(message.chat.id,
                     f"To continue, choose your language:", reply_markup=markup)


@bot.callback_query_handler(func=lambda call: call.data in languages)
def choose_lg(call):
    message = call.message
    if call.data not in languages:
        return
    if message.chat.id not in users:
        users[message.chat.id] = User(message.chat.id)
    users[message.chat.id].language = call.data
    with open(users[message.chat.id].language + '.json', encoding='utf-8') as f:
        msg = json.load(f)['welcome']
    markup = types.InlineKeyboardMarkup()
    b1 = types.InlineKeyboardButton(msg[1], callback_data='reg')
    b2 = types.InlineKeyboardButton(msg[2], callback_data='log')
    markup.row(b1)
    markup.row(b2)
    bot.edit_message_text(chat_id=message.chat.id, message_id=message.id, text=msg[0], reply_markup=markup)


@bot.callback_query_handler(func=lambda call: call.data == 'reg')
def reg(call):
    message = call.message
    with open(users[message.chat.id].language + '.json', encoding='utf-8') as f:
        jsf = json.load(f)
    print(wb.sheetnames, call.message.chat.id)
    if str(call.message.chat.id) in wb.sheetnames:
        markup = types.InlineKeyboardMarkup()
        b = types.InlineKeyboardButton(jsf['welcome'][2], callback_data='log')
        markup.row(b)
        bot.edit_message_text(chat_id=message.chat.id, message_id=message.id, text=jsf['linked_error'],
                              reply_markup=markup)
        return
    users[message.chat.id].message_id = message.id
    bot.edit_message_text(chat_id=message.chat.id, message_id=message.id, text=jsf['FIO'])
    bot.register_next_step_handler(call.message, get_name)


def get_name(message):
    with open(users[message.chat.id].language + '.json', encoding='utf-8') as f:
        jsf = json.load(f)
    name = message.text
    mid = users[message.chat.id].message_id
    print(name)
    if len(name.split()) != 2:
        try:
            bot.edit_message_text(chat_id=message.chat.id, message_id=mid, text=jsf['name_error'])
        except:
            pass
        bot.delete_message(chat_id=message.chat.id, message_id=message.id)
        bot.register_next_step_handler(message, get_name)
        return
    users[message.chat.id].name = name
    markup = types.InlineKeyboardMarkup()
    for key, item in jsf['gender'][1].items():
        b = types.InlineKeyboardButton(item, callback_data=key)
        markup.row(b)
    bot.delete_message(chat_id=message.chat.id, message_id=message.id)
    bot.edit_message_text(chat_id=message.chat.id, message_id=mid, text=jsf['gender'][0], reply_markup=markup)


@bot.callback_query_handler(func=lambda call: call.data in ['male', 'female', 'another'])
def get_gender(call):
    message = call.message
    gender = call.data
    users[message.chat.id].gender = gender
    with open(users[message.chat.id].language + '.json', encoding='utf-8') as f:
        jsf = json.load(f)
    markup = types.InlineKeyboardMarkup()
    for key, item in jsf['region'][1].items():
        b = types.InlineKeyboardButton(item, callback_data=key)
        markup.row(b)
    txt = jsf['region'][0]
    bot.edit_message_text(chat_id=message.chat.id, message_id=message.id, text=txt, reply_markup=markup)


@bot.callback_query_handler(func=lambda call: call.data in ['russia', 'europe', 'asia', 'australia', 'south america', 'nord america'])
def get_region(call):
    region = call.data
    users[call.message.chat.id].region = region
    with open(users[call.message.chat.id].language + '.json', encoding='utf-8') as f:
        jsf = json.load(f)
    bot.edit_message_text(chat_id=call.message.chat.id, message_id=call.message.id, text=jsf['date'])
    bot.register_next_step_handler(call.message, get_date)


def get_date(message):
    date = message.text.split()
    print(date)
    with open(users[message.chat.id].language + '.json', encoding='utf-8') as f:
        jsf = json.load(f)
    try:
        time_now = dtm.date.today()
        time = dtm.date(year=int(date[2]), month=int(date[1]), day=int(date[0]))
        if time > time_now:
            raise Exception
        if time_now.year - int(date[2]) > 170:
            raise Exception
    except:
        try:
            bot.edit_message_text(chat_id=message.chat.id, text=jsf['date_error'],
                                  message_id=users[message.chat.id].message_id)
        except:
            pass
        bot.delete_message(chat_id=message.chat.id, message_id=message.id)
        bot.register_next_step_handler(message, get_date)
        return
    users[message.chat.id].date = ' '.join(date)
    bot.delete_message(chat_id=message.chat.id, message_id=message.id)
    bot.edit_message_text(chat_id=message.chat.id, text=jsf['nickname'], message_id=users[message.chat.id].message_id)
    bot.register_next_step_handler(message, get_nickname)


def get_nickname(message):
    global wb
    nn = message.text
    with open(users[message.chat.id].language + '.json', encoding='utf-8') as f:
        jsf = json.load(f)
    nicks = []
    for i in wb.sheetnames:
        nicks.append(wb[i]['A2'].value)
    if nn in nicks:
        bot.delete_message(chat_id=message.chat.id, message_id=message.id)
        try:
            bot.edit_message_text(chat_id=message.chat.id, text=jsf['nick_error'],
                                  message_id=users[message.chat.id].message_id)
        except:
            pass
        bot.register_next_step_handler(message, get_nickname)
        return
    users[message.chat.id].nname = nn
    bot.delete_message(chat_id=message.chat.id, message_id=message.id)
    bot.edit_message_text(chat_id=message.chat.id, text=jsf['psw'], message_id=users[message.chat.id].message_id)
    bot.register_next_step_handler(message, create_password)


def create_password(message):
    psw = message.text
    with open(users[message.chat.id].language + '.json', encoding='utf-8') as f:
        jsf = json.load(f)
    if len(psw) < 6:
        try:
            bot.edit_message_text(chat_id=message.chat.id, text=jsf['psw_error'],
                                  message_id=users[message.chat.id].message_id)
        except:
            pass
        bot.delete_message(message.chat.id, message.id)
        bot.register_next_step_handler(message, create_password)
        return
    users[message.chat.id].tg_name = message.from_user.username
    users[message.chat.id].upload_info()
    passwords[message.chat.id] = psw
    print(users[message.chat.id].tg_name, 'зарегестрировался')
    with open("passwords.json", "w") as write_file:
        json.dump(passwords, write_file)
    markup = types.InlineKeyboardMarkup()
    for key, elem in jsf["choose_role"][1].items():
        b = types.InlineKeyboardButton(elem, callback_data=key)
        markup.row(b)
    bot.edit_message_text(chat_id=message.chat.id, text=jsf['choose_role'][0], reply_markup=markup,
                          message_id=users[message.chat.id].message_id)
    bot.delete_message(message.chat.id, message.id)


@bot.callback_query_handler(func=lambda call: call.data in ["stay_user", "become_beta", "become_dev"])
def get_role(call):
    with open(users[call.message.chat.id].language + '.json', encoding='utf-8') as f:
        jsf = json.load(f)
    if call.data == 'stay_user':
        markup = types.InlineKeyboardMarkup().row(types.InlineKeyboardButton(jsf['mm'], callback_data='main_menu'))
        bot.edit_message_text(chat_id=call.message.chat.id, text=jsf['suc_reg'], message_id=call.message.id,
                              reply_markup=markup)
    else:
        bot.edit_message_text(text=jsf["choose_role"][2], chat_id=call.message.chat.id, message_id=call.message.id)
        if call.data == 'become_beta':
            bot.register_next_step_handler(call.message, do_beta)
        else:
            bot.register_next_step_handler(call.message, do_dev)


def do_beta(message):
    with open(users[message.chat.id].language + '.json', encoding='utf-8') as f:
        jsf = json.load(f)
    markup = types.InlineKeyboardMarkup().row(types.InlineKeyboardButton(jsf['mm'], callback_data='main_menu'))
    bot.edit_message_text(chat_id=message.chat.id, text=jsf['application'],
                          message_id=users[message.chat.id].message_id, reply_markup=markup)
    bot.delete_message(chat_id=message.chat.id, message_id=message.id)
    for a in admins_id:
        bot.send_message(a, f'@{message.from_user.username} хочет быть бета-тестером со словами:\n{message.text}')


def do_dev(message):
    with open(users[message.chat.id].language + '.json', encoding='utf-8') as f:
        jsf = json.load(f)
    markup = types.InlineKeyboardMarkup().row(types.InlineKeyboardButton(jsf['mm'], callback_data='main_menu'))
    bot.edit_message_text(chat_id=message.chat.id, text=jsf['application'],
                          message_id=users[message.chat.id].message_id, reply_markup=markup)
    bot.delete_message(chat_id=message.chat.id, message_id=message.id)
    for a in admins_id:
        bot.send_message(a, f'@{message.from_user.username} хочет быть разрабом со словами:\n{message.text}')


@bot.callback_query_handler(func=lambda call: call.data == 'log')
def log(call):
    with open(users[call.message.chat.id].language + '.json', encoding='utf-8') as f:
        jsf = json.load(f)
    if str(call.message.chat.id) in wb.sheetnames:
        users[call.message.chat.id].load_info()
        print(users[call.message.chat.id].tg_name, 'вошел в аккаунт')
        markup = types.InlineKeyboardMarkup()
        markup.row(types.InlineKeyboardButton(jsf['mm'], callback_data='main_menu'))
        bot.edit_message_text(chat_id=call.message.chat.id, message_id=call.message.id,
                              text=jsf['log_suc'] + ' ' + users[call.message.chat.id].nname, reply_markup=markup)
    else:
        markup = types.InlineKeyboardMarkup()
        b = types.InlineKeyboardButton(jsf['welcome'][1], callback_data='reg')
        markup.row(b)
        bot.edit_message_text(chat_id=call.message.chat.id, message_id=call.message.id,
                              text=jsf['login_error'], reply_markup=markup)


@bot.message_handler(content_types=['text'])
def deleter(message):
    bot.delete_message(message.chat.id, message.id)


@bot.callback_query_handler(func=lambda call: call.data == 'main_menu')
def main_menu(call):
    message = call.message
    with open(users[message.chat.id].language + '.json', encoding='utf-8') as f:
        jsf = json.load(f)
    markup = types.InlineKeyboardMarkup()
    for key, elem in jsf['main_menu'][1].items():
        b = types.InlineKeyboardButton(elem, callback_data=key)
        markup.row(b)
    bot.edit_message_text(chat_id=message.chat.id, message_id=call.message.id,
                          text=jsf['main_menu'][0], reply_markup=markup)


@bot.callback_query_handler(func=lambda call: call.data == 'news')
def news(call):
    message = call.message
    with open(users[message.chat.id].language + '.json', encoding='utf-8') as f:
        jsf = json.load(f)
    with open(f'news_{users[message.chat.id].language}.txt', encoding='utf-8') as f:
        active_news = f.read()
    if not active_news:
        active_news = [jsf['news_events'][0]]
    else:
        active_news = active_news.split('***')
    markup = types.InlineKeyboardMarkup()
    b = []
    for key, elem in jsf['news_events'][-1].items():
        b.append(types.InlineKeyboardButton(elem, callback_data=key))
    markup.add(*b )
    markup.row(types.InlineKeyboardButton(jsf['mm'], callback_data='main_menu'))
    users[message.chat.id].news_day = 0
    bot.edit_message_text(chat_id=message.chat.id, message_id=message.id, text=active_news[0],
                          reply_markup=markup)


@bot.callback_query_handler(func=lambda hui: hui.data == 'settings')
def settings(call):
    message = call.message
    with open(users[message.chat.id].language + '.json', encoding='utf-8') as f:
        jsf = json.load(f)
    markup = types.InlineKeyboardMarkup()
    for key, elem in jsf['settings'][1].items():
        b = types.InlineKeyboardButton(elem, callback_data=key)
        markup.row(b)
    markup.row(types.InlineKeyboardButton(jsf['mm'], callback_data='main_menu'))
    bot.edit_message_text(chat_id=message.chat.id, message_id=message.id, text=jsf['settings'][0], reply_markup=markup)


@bot.callback_query_handler(func=lambda call: call.data == 'choose_lg')
def rechoose_lg(call):
    message = call.message
    with open(users[message.chat.id].language + '.json', encoding='utf-8') as f:
        jsf = json.load(f)
    markup = types.InlineKeyboardMarkup()
    for lg in languages:
        b1 = types.InlineKeyboardButton(languages_emojies[lg] + lg + languages_emojies[lg], callback_data=lg + 'n')
        markup.row(b1)
    bot.edit_message_text(chat_id=message.chat.id, message_id=message.id, text=jsf['rechoce'], reply_markup=markup)


@bot.callback_query_handler(func=lambda call: call.data[:-1] in languages)
def confrim_lgc(call):
    message = call.message
    users[message.chat.id].language = call.data[:-1]
    with open(users[message.chat.id].language + '.json', encoding='utf-8') as f:
        jsf = json.load(f)
    markup = types.InlineKeyboardMarkup()
    markup.row(types.InlineKeyboardButton(jsf['mm'], callback_data='main_menu'))
    bot.edit_message_text(chat_id=message.chat.id, message_id=message.id, text=jsf['suc_rec'], reply_markup=markup)


@bot.callback_query_handler(func=lambda amd_sosat: amd_sosat.data == 'del_ac')
def del_ac(call):
    message = call.message
    with open(users[message.chat.id].language + '.json', encoding='utf-8') as f:
        jsf = json.load(f)
    markup = types.InlineKeyboardMarkup()
    for key, elem in jsf['del_conf'][1].items():
        b = types.InlineKeyboardButton(elem, callback_data=key)
        markup.row(b)
    bot.edit_message_text(chat_id=message.chat.id, message_id=message.id, text=jsf['del_conf'][0], reply_markup=markup)


@bot.callback_query_handler(func=lambda call: call.data == "conf")
def del_conf(call):
    message = call.message
    with open(users[message.chat.id].language + '.json', encoding='utf-8') as f:
        jsf = json.load(f)
    users[call.message.chat.id] = None
    wb.remove(wb[str(call.message.chat.id)])
    wb.save('data.xlsx')
    bot.edit_message_text(chat_id=message.chat.id, message_id=message.id, text=jsf['suc_del'])


@bot.callback_query_handler(func=lambda anime_for_gays: anime_for_gays.data == 'forg_pass')
def forg_pass(call):
    message = call.message
    with open(users[message.chat.id].language + '.json', encoding='utf-8') as f:
        jsf = json.load(f)
    with open('passwords.json', encoding='utf-8') as f:
        data = json.load(f)
    markup = types.InlineKeyboardMarkup().row(types.InlineKeyboardButton(jsf['del_conf'][1]['settings'],
                                                                         callback_data='settings'))
    print(data)
    bot.edit_message_text(chat_id=message.chat.id, message_id=message.id,
                          text=jsf['rem_pas'] + data[str(message.chat.id)], reply_markup=markup)


@bot.callback_query_handler(func=lambda call: call.data in ["n_back", 'n_forw'])
def news_engine(call):
    with open(users[call.message.chat.id].language + '.json', encoding='utf-8') as f:
        jsf = json.load(f)
    with open(f'news_{users[call.message.chat.id].language}.txt', encoding='utf-8') as f:
        active_news = f.read()
    if not active_news:
        active_news = [jsf['news_events'][0]]
    else:
        active_news = active_news.split('***')
    if call.data == 'n_back':
        if len(active_news) - 1 == users[call.message.chat.id].news_day:
            bot.answer_callback_query(callback_query_id=call.id, text=jsf['news_events'][1])
            return
        users[call.message.chat.id].news_day += 1
    elif call.data == 'n_forw':
        if users[call.message.chat.id].news_day == 0:
            bot.answer_callback_query(callback_query_id=call.id, text=jsf['news_events'][1])
            return
        users[call.message.chat.id].news_day -= 1
    markup = types.InlineKeyboardMarkup()
    b = []
    for key, elem in jsf['news_events'][-1].items():
        b.append(types.InlineKeyboardButton(elem, callback_data=key))
    markup.add(*b)
    markup.row(types.InlineKeyboardButton(jsf['mm'], callback_data='main_menu'))
    bot.edit_message_text(chat_id=call.message.chat.id, message_id=call.message.id,
                          text=active_news[users[call.message.chat.id].news_day], reply_markup=markup)


bot.remove_webhook()
bot.infinity_polling()
